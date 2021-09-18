#!/usr/bin/env python
import os
import sys
import glob
import shutil
from re import sub
from datetime import date, datetime
from copy import copy
import warnings


def main():
    try:
        #provide the folder that contains all the unprocessed, input files needed for the script
        folder = os.path.normpath(sys.argv[1])
        if os.path.isfile(folder):
            folder = os.path.dirname(folder)
            print(f'INFO: A path to a file was provided. Will attempt to use the following folder path instead: {folder}')
    except IndexError:
        folder = 'UVR_Files'
        print(f'INFO: No folder provided, will default to folder "{folder}"')

    if not os.path.isdir(folder):
        sys.exit(f'ERROR: There is no folder named "{folder}" to read from. Program will exit.')

    import_required_modules()

    output_folder = os.path.join(folder, 'processed_files')
    if os.path.isdir(output_folder):
        shutil.rmtree(output_folder)
        os.makedirs(output_folder)
    else:
        os.makedirs(output_folder)

    monthyear = get_month_and_year()
    print('INFO: File processing may take up to 2 minutes...')

    process_it_ams_access_file(folder, output_folder, monthyear)
    process_ogm_file(folder, output_folder, monthyear)
    process_regional_files(folder, output_folder, monthyear)
    process_pod_file(folder, output_folder, monthyear)
    process_tta_file(folder, output_folder, monthyear)
    process_monitoring_file(folder, output_folder, monthyear)
    print('FINISHED')


def import_required_modules():
    from subprocess import check_call
    global pd, np, Font, PatternFill, Border, Side, Alignment, DataValidation, load_workbook, relativedelta
    try:
        import pandas as pd
    except ImportError:
        print('INFO: The script requires the python library "pandas" to be installed. Attempting to install now:')
        try:
            check_call([sys.executable, '-m', 'pip', 'install', 'pandas'])
            import pandas as pd
            print('INFO: Successfully installed and imported "pandas"')
        except:
            try:
                check_call([sys.executable, '-m', 'ensurepip', '--upgrade'])
                check_call([sys.executable, '-m', 'pip', 'install', 'pandas'])
                import pandas as pd
                print('INFO: Successfully installed and imported "pandas"')
            except:
                sys.exit('ERROR: Failed to install "pandas". Program will exit. Please ensure you have pip installed and install the library yourself manually in the terminal: enter "pip3 install pandas" or "pip install pandas"')
    
    try:
        import numpy as np
    except ImportError:
        print('INFO: The script requires the python library "numpy" to be installed. Attempting to install now:')
        try:
            check_call([sys.executable, '-m', 'pip', 'install', 'numpy'])
            import numpy as np
            print('INFO: Successfully installed and imported "numpy"')
        except:
            sys.exit('ERROR: Failed to install "numpy". Program will exit. Please ensure you have pip installed and install the library yourself manually in the terminal: enter "pip3 install numpy" or "pip install numpy"')
    
    try:
        from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
        from openpyxl.worksheet.datavalidation import DataValidation
        from openpyxl import load_workbook
    except ImportError:
        print('INFO: The script requires the python library "openpyxl" to be installed. Attempting to install now:')
        try:
            check_call([sys.executable, '-m', 'pip', 'install', 'openpyxl'])
            from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
            from openpyxl.worksheet.datavalidation import DataValidation
            from openpyxl import load_workbook
            print('INFO: Successfully installed and imported "openpyxl"')
        except:
            sys.exit('ERROR: Failed to install "openpyxl". Program will exit. Please ensure you have pip installed and install the library yourself manually in the terminal: enter "pip3 install openpyxl" or "pip install openpyxl"')

    try:
        from dateutil.relativedelta import relativedelta
    except ImportError:
        print('INFO: The script requires the python library "dateutil" to be installed. Attempting to install now:')
        try:
            check_call([sys.executable, '-m', 'pip', 'install', 'python-dateutil'])
            from dateutil.relativedelta import relativedelta
            print('INFO: Successfully installed and imported "dateutil"')
        except:
            sys.exit('ERROR: Failed to install "dateutil". Program will exit. Please ensure you have pip installed and install the library yourself manually in the terminal: enter "pip3 install python-dateutil" or "pip install python-dateutil"')


def process_ogm_file(input_folder, output_folder, monthyear):
    final_ogm_filepath = os.path.join(output_folder, f'HSES OGM Accounts_{monthyear}.xlsx')
    processed_it_ams_filepath = os.path.join(output_folder, f'IT-AMS Access_{monthyear}.xlsx')

    #glob.glob returns an array of matching filenames, we use this to check if the file exists and to get the filepath
    rgnall = glob.glob(os.path.join(input_folder, 'RgnAll HSES Accounts*.xlsx'))
    rgnall_file_exists = True if len(rgnall) > 0 else False
    rgn0 = glob.glob(os.path.join(input_folder, 'Rgn0 OGM Accounts*.xlsx'))
    rgn0_file_exists = True if len(rgn0) > 0 else False
    role = glob.glob(os.path.join(input_folder, 'UserRoleListingReport*.xlsx'))
    role_file_exists = True if len(role) > 0 else False
    ogm_filepath = os.path.join(input_folder, f'HSES OGM Accounts_{monthyear}.xlsx')
    ogm_file_already_exists = True if len(glob.glob(ogm_filepath)) > 0 else False

    if not ogm_file_already_exists and rgnall_file_exists and rgn0_file_exists and role_file_exists and os.path.isfile(processed_it_ams_filepath):
        rgnall_filepath = rgnall[0]
        rgn0_filepath = rgn0[0]
        role_filepath = role[0]
        # copy file over to the processed_files folder, this will be our ogm file after it has been processed
        shutil.copy(rgnall_filepath, final_ogm_filepath)

        # this will get rid of the unnessecary warnings in the log
        with warnings.catch_warnings(record=True):
            warnings.simplefilter("always")

            ogm_df = pd.read_excel(final_ogm_filepath)
            ogm_df = ogm_df[
                (ogm_df['Roles'].str.lower().str.contains('user verification contact-program') == False) &
                (
                    (ogm_df['Roles'].str.lower().str.contains('grants management officer')) |
                    (ogm_df['Roles'].str.lower().str.contains('grants specialist')) |
                    (ogm_df['Roles'].str.lower().str.contains('grants admin support'))
                )
            ]

            central_office_df = pd.read_excel(rgn0_filepath)
            central_office_df['Region'] = central_office_df['Region'].str.replace('Central Office', '0')
            ogm_df = pd.concat([ogm_df, central_office_df], axis=0)

            it_ams_df = pd.read_excel(processed_it_ams_filepath)
            # drop and re-add IT-AMS Access column from IT-AMS Access file
            ogm_df = ogm_df.iloc[:,:-1]
            ogm_df = pd.merge(ogm_df, it_ams_df[['Email', 'IT-AMS Access']], how='left', on='Email')

            user_role_df = pd.read_excel(role_filepath)
            ogm_df = pd.merge(ogm_df, user_role_df[['Email', 'User Location']], how='left', on='Email')
            if user_role_df['User Location'].dtype == 'int64' and ogm_df['User Location'].dtype == 'float64':
                ogm_df['User Location'] = ogm_df['User Location'].astype('Int64')
            rearrange_cols = list(ogm_df.columns)
            rearrange_cols.pop()
            rearrange_cols.insert(1, 'User Location')
            ogm_df = ogm_df[rearrange_cols]
            ogm_df = ogm_df.rename(columns={'IT-AMS Access': 'IT-AMS Role\n(please specify using dropdown)'})
            ogm_df.loc[ogm_df['User Location'].isna(), 'User Location'] = 0
            ogm_df = ogm_df.sort_values(by=['User Location', 'Last Name', 'First Name'])

            ogm_df.to_excel(final_ogm_filepath, 'OGM HSES Accounts', index=False)
            wb = load_workbook(final_ogm_filepath)
            ws = wb.active
            style_worksheet(ws)
            separate_location_groups_with_thick_borders(ws)
            add_it_ams_roles_sheet(wb)
            wb.save(final_ogm_filepath)
            print(f'File processed: {final_ogm_filepath}')
    else:
        if ogm_file_already_exists:
            print('INFO: HSES OGM Accounts_<month>-<year>.xlsx already exists in the folder and is therefore assumed to be intentionally provided. No action will be taken to process this file, it will be used as is and copied over to the output folder.')
            shutil.copy(ogm_filepath, final_ogm_filepath)
        else:
            print('FAILED: There are one or more files missing needed to generate the HSES OGM Accounts report.')
            print('Make sure you provided the correct files/file name formats and/or the correct folder/directory path.')
            if not rgnall_file_exists:
                print('Missing file: RgnAll HSES Accounts.xlsx')
            if not rgn0_file_exists:
                print('Missing file: Rgn0 OGM Accounts.xlsx (Central Office OGM Accounts)')
            if not role_file_exists:
                print('Missing file: UserRoleListingReport.xlsx') 
            if not os.path.isfile(processed_it_ams_filepath):
                print('Missing file: IT-AMS Access_<month>-<year>.xlsx')


def process_regional_files(input_folder, output_folder, monthyear):
    processed_ogm_filepath = os.path.join(output_folder, f'HSES OGM Accounts_{monthyear}.xlsx')
    processed_it_ams_filepath = os.path.join(output_folder, f'IT-AMS Access_{monthyear}.xlsx')
    regional_files_list = glob.glob(os.path.join(input_folder, 'Rgn[0-9][0-9]*'))
    if os.path.isfile(processed_ogm_filepath) and os.path.isfile(processed_it_ams_filepath) and len(regional_files_list) > 0:
        regional_files_list.sort()
        if len(regional_files_list) < 12:
            print('WARNING: Less than 12 Regional files were provided/detected. There should be 12 of these files (Rgn<##> HSES Accounts.xlsx). Please verify')
        for region in regional_files_list:
            final_region_filepath = os.path.join(output_folder, os.path.basename(region))
            final_region_filepath = sub(r'(\s*\(\d*\)\.xlsx)|(\.xlsx)', f'_{monthyear}.xlsx', final_region_filepath)
            shutil.copy(region, final_region_filepath)
            with warnings.catch_warnings(record=True):
                warnings.simplefilter("always")
                xl = pd.ExcelFile(final_region_filepath)
                region_df = pd.read_excel(xl, 0)

                ogm_df = pd.read_excel(processed_ogm_filepath)
                region_df = region_df[~region_df['Email'].isin(ogm_df['Email'].tolist())]

                it_ams_df = pd.read_excel(processed_it_ams_filepath)
                # drop and re-add IT-AMS Access column from IT-AMS Access file
                region_df = region_df.iloc[:,:-1]
                region_df = pd.merge(region_df, it_ams_df[['Email', 'IT-AMS Access']], how='left', on='Email')
                region_df = region_df.rename(columns={'IT-AMS Access': 'IT-AMS Role\n(please specify using dropdown)'})
                region_df = region_df.sort_values(by=['Last Name', 'First Name'])

                writer = pd.ExcelWriter(final_region_filepath)
                region_df.to_excel(writer, xl.sheet_names[0], index=False)
                if xl.sheet_names[1]:
                    sheet2 = pd.read_excel(region, xl.sheet_names[1]) # get it from original file to avoid write errors
                    sheet2.to_excel(writer, xl.sheet_names[1], index=False)
                writer.save()
                wb_region = load_workbook(final_region_filepath)
                ws_region = wb_region.active
                style_worksheet(ws_region)
                sheet2 = wb_region[xl.sheet_names[1]]
                style_worksheet(sheet2)
                add_it_ams_roles_sheet(wb_region)
                wb_region.save(final_region_filepath)
                print(f'File processed: {final_region_filepath}')
    else:
        print('FAILED: There are one or more files missing needed to process Rgn<##> HSES Accounts_<date>.xlsx files.')
        if len(regional_files_list) == 0:
            print('Missing file(s): Rgn<##> HSES Accounts.xlsx')
        if not os.path.isfile(processed_ogm_filepath):
            print('Missing file: HSES OGM Accounts_<month>-<year>.xlsx')
        if not os.path.isfile(processed_it_ams_filepath):
            print('Missing file: IT-AMS Access_<month>-<year>.xlsx')


def process_it_ams_access_file(input_folder, output_folder, monthyear):
    final_it_ams_filepath = os.path.join(output_folder, f'IT-AMS Access_{monthyear}.xlsx')

    rgnall = glob.glob(os.path.join(input_folder, 'RgnAll HSES Accounts*.xlsx'))
    rgnall_file_exists = True if len(rgnall) > 0 else False
    rgn0 = glob.glob(os.path.join(input_folder, 'Rgn0 OGM Accounts*.xlsx'))
    rgn0_file_exists = True if len(rgn0) > 0 else False
    rgn0_pod = glob.glob(os.path.join(input_folder, 'Rgn0 HSES POD Accounts*.xlsx'))
    rgn0_pod_file_exists = True if len(rgn0_pod) > 0 else False
    it_ams_filepath = os.path.join(input_folder, f'IT-AMS Access_{monthyear}.xlsx')
    it_ams_file_already_exists = True if len(glob.glob(it_ams_filepath)) > 0 else False

    if not it_ams_file_already_exists and rgnall_file_exists and rgn0_pod_file_exists:
        rgnall_filepath = rgnall[0]
        rgn0_filepath = rgn0[0]
        rgn0_pod_filepath = rgn0_pod[0]
        shutil.copy(rgnall_filepath, final_it_ams_filepath)

        with warnings.catch_warnings(record=True):
            warnings.simplefilter("always")

            it_ams_df = pd.read_excel(final_it_ams_filepath)

            central_office_df = pd.read_excel(rgn0_filepath)
            central_office_df['Region'] = central_office_df['Region'].str.replace('Central Office', '0')
            it_ams_df = pd.concat([it_ams_df, central_office_df], axis=0)

            central_office_pod_df = pd.read_excel(rgn0_pod_filepath)
            central_office_pod_df['Region'] = central_office_pod_df['Region'].str.replace('Central Office', '0')
            it_ams_df = pd.concat([it_ams_df, central_office_pod_df], axis=0)

            it_ams_df.insert(loc=len(it_ams_df.columns) - 1, column='RPM', value=np.nan)
            it_ams_df.insert(loc=len(it_ams_df.columns) - 1, column='PS', value=np.nan)
            it_ams_df.insert(loc=len(it_ams_df.columns) - 1, column='GS', value=np.nan)
            it_ams_df.insert(loc=len(it_ams_df.columns) - 1, column='SPS', value=np.nan)

            it_ams_df.loc[it_ams_df['Roles'].str.contains('IT-AMS RPM Application Access'), 'RPM'] = 'RPM'
            it_ams_df.loc[it_ams_df['Roles'].str.contains('Supervisory Program Specialist'), 'SPS'] = 'SPS'
            it_ams_df.loc[(it_ams_df['Roles'].str.contains('IT-AMS PS Application Access')) |
                ((it_ams_df['Roles'].str.contains('Program Specialist')) & 
                (~it_ams_df['Roles'].str.contains('Supervisory Program Specialist'))), 'PS'] = 'PS'
            it_ams_df.loc[(it_ams_df['Roles'].str.contains('IT-AMS GS Application Access')) |
                ((it_ams_df['Roles'].str.contains('Grants Specialist')) & 
                (~it_ams_df['Roles'].str.contains('National Centers Grants Specialist'))), 'GS'] = 'GS'

            # now fill in the IT-AMS Access column based on the above columns
            it_ams_df.loc[(it_ams_df['RPM'].str.contains('RPM')) & 
                (it_ams_df['PS'].isna()) & (it_ams_df['GS'].isna()) &
                (it_ams_df['SPS'].isna()), 'IT-AMS Access'] = 'RPM'
            it_ams_df.loc[(it_ams_df['SPS'].str.contains('SPS')) & 
                (it_ams_df['PS'].isna()) & (it_ams_df['GS'].isna()) &
                (it_ams_df['RPM'].isna()), 'IT-AMS Access'] = 'SPS'
            it_ams_df.loc[(it_ams_df['PS'].str.contains('PS')) & 
                (it_ams_df['SPS'].isna()) & (it_ams_df['GS'].isna()) &
                (it_ams_df['RPM'].isna()), 'IT-AMS Access'] = 'PS'
            it_ams_df.loc[(it_ams_df['GS'].str.contains('GS')) & 
                (it_ams_df['SPS'].isna()) & (it_ams_df['PS'].isna()) &
                (it_ams_df['RPM'].isna()), 'IT-AMS Access'] = 'GS'
            it_ams_df.loc[(it_ams_df['PS'].str.contains('PS')) & 
                (it_ams_df['SPS'].isna()) & (it_ams_df['GS'].str.contains('GS')) &
                (it_ams_df['RPM'].isna()), 'IT-AMS Access'] = 'PS and GS'
                
            # need custom sorting as the Region column contains both integers (one region) and a string list of integers (multiple regions: '0,1,4,5')
            sort_on_one_region = it_ams_df[~it_ams_df['Region'].str.contains(',')]
            sort_on_one_region['Region'] = sort_on_one_region['Region'].astype('float').astype('Int64')
            sort_on_one_region = sort_on_one_region.sort_values(by=['Region', 'Last Name', 'First Name'])
            sort_on_one_region['Region'] = sort_on_one_region['Region'].astype('object')
            sort_on_multiple_regions = it_ams_df[it_ams_df['Region'].str.contains(',')].sort_values(by=['Region', 'Last Name', 'First Name'])
            it_ams_df = pd.concat([sort_on_one_region, sort_on_multiple_regions], axis=0)

            it_ams_df.to_excel(final_it_ams_filepath, 'IT-AMS Roles', index=False)
            wb = load_workbook(final_it_ams_filepath)
            ws = wb.active
            style_worksheet(ws)
            wb.save(final_it_ams_filepath)
            print(f'File processed: {final_it_ams_filepath}')
    else:
        if it_ams_file_already_exists:
            print('INFO: IT-AMS Access_<month>-<year>.xlsx already exists in the folder and is therefore assumed to be intentionally provided. No action will be taken to process this file, it will be used as is and copied over to the output folder.')
            shutil.copy(it_ams_filepath, final_it_ams_filepath)
        else:
            print('FAILED: There are one or more files missing needed to generate the IT-AMS Access report.')
            print('Make sure you provided the correct files/file name formats and/or the correct folder/directory path.')
            if not rgnall_file_exists:
                print('Missing file: RgnAll HSES Accounts.xlsx')
            if not rgn0_file_exists:
                print('Missing file: Rgn0 OGM Accounts.xlsx (Central Office OGM Accounts)')
            if not rgn0_pod_file_exists:
                print('Missing file: Rgn0 HSES POD Accounts.xlsx (Central Office POD Accounts)') 


def process_monitoring_file(input_folder, output_folder, monthyear):
    final_monitoring_filepath = os.path.join(output_folder, f'HSES Monitoring Network Accounts_{monthyear}.xlsx')

    danya = glob.glob(os.path.join(input_folder, 'Danya User HSES Accounts*.xlsx'))
    danya_file_exists = True if len(danya) > 0 else False
    lewin = glob.glob(os.path.join(input_folder, 'Lewin Accounts*.xlsx'))
    lewin_file_exists = True if len(lewin) > 0 else False
    network_users = glob.glob(os.path.join(input_folder, 'Monitoring_Network_Users*.xlsx'))
    network_users_file_exists = True if len(network_users) > 0 else False

    if danya_file_exists and lewin_file_exists and network_users_file_exists:
        danya_filepath = danya[0]
        lewin_filepath = lewin[0]
        network_users_filepath = network_users[0]
        with warnings.catch_warnings(record=True):
            warnings.simplefilter("always")
            support_accounts_df = pd.read_excel(danya_filepath, 'Verify Review Support Accounts')
            planner_accounts_df = pd.read_excel(danya_filepath, 'Verify Review Planner Accounts')
            support_accounts_df = pd.concat([support_accounts_df, planner_accounts_df], axis=0)

            lewin_df = pd.read_excel(lewin_filepath)
            copy_to_lewin = support_accounts_df[support_accounts_df['Roles'].str.contains('Lewin Group')]
            # swap the Title and Roles columns as they are out of order when copied over
            cols_list = list(copy_to_lewin.columns)
            roles_col = cols_list.pop()
            title_col = cols_list.pop()
            cols_list += [roles_col, title_col]
            copy_to_lewin = copy_to_lewin[cols_list]
            lewin_df = pd.concat([lewin_df, copy_to_lewin], axis=0)
            lewin_df = lewin_df.sort_values(by=['Last Name', 'First Name'])
            support_accounts_df = support_accounts_df[~support_accounts_df['Roles'].str.contains('Lewin Group')]
            support_accounts_df = support_accounts_df.sort_values(by=['Title', 'Last Name', 'First Name'])

            reviewer_accounts_df = pd.read_excel(danya_filepath, 'Verify Reviewer Accounts')
            network_users_df = pd.read_excel(network_users_filepath)
            reviewer_accounts_df = pd.merge(reviewer_accounts_df, network_users_df[['Email', 'Gateway Id']], how='left', on='Email')
            reviewer_accounts_df = reviewer_accounts_df.rename(columns={'Gateway Id': 'Monitoring System ID Linked for Reviews'})
            reviewer_accounts_df = reviewer_accounts_df.sort_values(by=['Last Name', 'First Name'])
            # move users with no id to the top
            reviewer_accounts_df.loc[
                reviewer_accounts_df['Monitoring System ID Linked for Reviews'] == '0', 
                'Monitoring System ID Linked for Reviews'
                ] = np.nan
            reviewer_accounts_df = pd.concat([
                reviewer_accounts_df[reviewer_accounts_df['Monitoring System ID Linked for Reviews'].isna()],
                reviewer_accounts_df[~reviewer_accounts_df['Monitoring System ID Linked for Reviews'].isna()]
                ])

            writer = pd.ExcelWriter(final_monitoring_filepath)
            support_accounts_df.to_excel(writer, 'Verify Planner-Support Accounts', index=False)
            reviewer_accounts_df.to_excel(writer, 'Verify Reviewer Accounts', index=False)
            lewin_df.to_excel(writer, 'Verify Lewin Accounts', index=False)
            writer.save()
            wb = load_workbook(final_monitoring_filepath)
            ws = wb.active
            style_worksheet(ws)
            separate_title_groups_with_thick_borders(ws)
            style_worksheet(wb['Verify Reviewer Accounts'])
            highlight_reviewer_accounts_with_no_id_yellow(wb['Verify Reviewer Accounts'])
            style_worksheet(wb['Verify Lewin Accounts'])
            wb.save(final_monitoring_filepath)
            print(f'File processed: {final_monitoring_filepath}')
    else:
        print('FAILED: There are one or more files missing needed to generate the Monitoring report.')
        if not danya_file_exists:
            print('Missing file: Danya User HSES Accounts.xlsx')
        if not lewin_file_exists:
            print('Missing file: Lewin Accounts.xlsx')
        if not network_users_file_exists:
            print('Missing file: Monitoring_Network_Users.xlsx')


def process_pod_file(input_folder, output_folder, monthyear):
    final_pod_filepath = os.path.join(output_folder, f'Rgn0 HSES POD Accounts_{monthyear}.xlsx')
    processed_ogm_filepath = os.path.join(output_folder, f'HSES OGM Accounts_{monthyear}.xlsx')
    processed_it_ams_filepath = os.path.join(output_folder, f'IT-AMS Access_{monthyear}.xlsx')

    rgn0_pod = glob.glob(os.path.join(input_folder, 'Rgn0 HSES POD Accounts*.xlsx'))
    rgn0_pod_file_exists = True if len(rgn0_pod) > 0 else False

    if rgn0_pod_file_exists and os.path.isfile(processed_ogm_filepath) and os.path.isfile(processed_it_ams_filepath):
        pod_filepath = rgn0_pod[0]
        shutil.copy(pod_filepath, final_pod_filepath)
        with warnings.catch_warnings(record=True):
            warnings.simplefilter("always")
            pod_df = pd.read_excel(final_pod_filepath)

            ogm_df = pd.read_excel(processed_ogm_filepath)
            pod_df = pod_df[~pod_df['Email'].isin(ogm_df['Email'].tolist())]

            it_ams_df = pd.read_excel(processed_it_ams_filepath)
            # drop and re-add IT-AMS Access column from IT-AMS Access file
            pod_df = pod_df.iloc[:,:-1]
            pod_df = pd.merge(pod_df, it_ams_df[['Email', 'IT-AMS Access']], how='left', on='Email')
            pod_df = pod_df.rename(columns={'IT-AMS Access': 'IT-AMS Role\n(please specify using dropdown)'})
            pod_df = pod_df.sort_values(by=['Last Name', 'First Name'])

            pod_df.to_excel(final_pod_filepath, 'Rgn0 HSES POD Accounts', index=False)
            wb = load_workbook(final_pod_filepath)
            ws = wb.active
            style_worksheet(ws)
            add_it_ams_roles_sheet(wb)
            wb.save(final_pod_filepath)
            print(f'File processed: {final_pod_filepath}')
    else:
        print('FAILED: There are one or more files missing needed to generate the CO POD Accounts report.')
        if not rgn0_pod_file_exists:
            print('Missing file: Rgn0 HSES POD Accounts.xlsx (Central Office POD Accounts)')
        if not os.path.isfile(processed_ogm_filepath):
            print('Missing file: HSES OGM Accounts_<month>-<year>.xlsx')
        if not os.path.isfile(processed_it_ams_filepath):
            print('Missing file: IT-AMS Access_<month>-<year>.xlsx')


def process_tta_file(input_folder, output_folder, monthyear):
    final_tta_filepath = os.path.join(output_folder, f'Rgn0 HSES T&TA Accounts_{monthyear}.xlsx')
    processed_ogm_filepath = os.path.join(output_folder, f'HSES OGM Accounts_{monthyear}.xlsx')

    rgn0_tta = glob.glob(os.path.join(input_folder, 'Rgn0 HSES T&TA Accounts*.xlsx'))
    rgn0_tta_file_exists = True if len(rgn0_tta) > 0 else False

    if rgn0_tta_file_exists and os.path.isfile(processed_ogm_filepath):
        tta_filepath = rgn0_tta[0]
        shutil.copy(tta_filepath, final_tta_filepath)
        with warnings.catch_warnings(record=True):
            warnings.simplefilter("always")
            tta_df = pd.read_excel(final_tta_filepath)
            tta_df = tta_df.loc[:, ~tta_df.columns.str.contains('^Unnamed')] # drop the empty column at the end that is there for some reason, remove this line if it is no longer there

            ogm_df = pd.read_excel(processed_ogm_filepath)
            tta_df = tta_df[~tta_df['Email'].isin(ogm_df['Email'].tolist())]

            tta_df.to_excel(final_tta_filepath, 'Rgn0 HSES T&TA Accounts', index=False)
            wb = load_workbook(final_tta_filepath)
            ws = wb.active
            style_worksheet(ws)
            wb.save(final_tta_filepath)
            print(f'File processed: {final_tta_filepath}')
    else:
        print('FAILED: There are one or more files missing needed to generate the CO TTA Accounts report.')
        if not rgn0_tta_file_exists:
            print('Missing file: Rgn0 HSES T&TA Accounts.xlsx (Central Office T&TA Accounts)')
        if not os.path.isfile(processed_ogm_filepath):
            print('Missing file: HSES OGM Accounts_<month>-<year>.xlsx')


def style_worksheet(ws):
    ws.freeze_panes = "B2"
    ws.auto_filter.ref = ws.dimensions
    for row in range(1, ws.max_row + 1):
        for col in range(1, ws.max_column + 1):
            #default style for all cells
            cell = ws.cell(row = row, column = col)
            cell.alignment = Alignment(wrap_text=True, vertical='top', horizontal='left')
            cell.border = Border(left=Side(border_style='thin'), right=Side(border_style='thin'), 
                                top=Side(border_style='thin'), bottom=Side(border_style='thin'))
            cell.font = Font(name='Arial', size=10, bold=False)
            if row == 1:
                #style header row
                cell.font= Font(name='Arial', size=10, bold=True)
                cell.border = Border(left=Side(border_style=cell.border.left.style), right=Side(border_style=cell.border.right.style), 
                                top=Side(border_style=cell.border.top.style), bottom=Side(border_style='thick'))
                ws.row_dimensions[row].height = 45
                ws.column_dimensions[cell.column_letter].width = 14.84 #column width is 14 in excel doc, need to add 0.84 to that to get the intended value with openpyxl
                cell.fill = PatternFill('solid', fgColor='CCFFCC') #default light green
                if cell.value in ['Action Required', 'IT-AMS Role']:
                    cell.fill = PatternFill('solid', fgColor='FFFF00') #color yellow
                elif cell.value == 'IT-AMS Role\n(please specify using dropdown)':
                    dv = DataValidation(type="list", formula1='"PS, GS, PS and GS, SPS, RPM"', allow_blank=True)
                    ws.add_data_validation(dv)
                    dv.add(f'{cell.column_letter}2:{cell.column_letter}{ws.max_row}')
                    cell.fill = PatternFill('solid', fgColor='FFFF00') #color yellow
                elif cell.value in ['RPM', 'PS', 'GS', 'SPS']:
                    cell.fill = PatternFill('solid', fgColor='FFFF00') #color yellow
                    ws.column_dimensions[cell.column_letter].width = 8.84
                elif cell.value == 'IT-AMS Access':
                    cell.fill = PatternFill('solid', fgColor='99CCFF') #color blue
                    ws.column_dimensions[cell.column_letter].width = 8.84
                elif cell.value == 'Monitoring System ID Linked for Reviews':
                    cell.fill = PatternFill('solid', fgColor='99CCFF') #color blue
                    ws.column_dimensions[cell.column_letter].width = 31.17
                elif cell.value == 'Roles':
                    ws.column_dimensions[cell.column_letter].width = 37.17
                elif cell.value == 'Grantee Name':
                    ws.column_dimensions[cell.column_letter].width = 37.17
                elif cell.value in ['Email', 'Email Address']:
                    ws.column_dimensions[cell.column_letter].width = 21.51
                elif cell.value == 'Title':
                    ws.column_dimensions[cell.column_letter].width = 21.51
            #apply bold borders to whole table outline
            if col == 1:
                cell.border = Border(left=Side(border_style='thick'), right=Side(border_style=cell.border.right.style), 
                                top=Side(border_style=cell.border.top.style), bottom=Side(border_style=cell.border.bottom.style))
            if col == ws.max_column:
                cell.border = Border(left=Side(border_style=cell.border.left.style), right=Side(border_style='thick'), 
                                top=Side(border_style=cell.border.top.style), bottom=Side(border_style=cell.border.bottom.style))
            if row == 1:
                cell.border = Border(left=Side(border_style=cell.border.left.style), right=Side(border_style=cell.border.right.style), 
                                top=Side(border_style='thick'), bottom=Side(border_style=cell.border.bottom.style))
            if row == ws.max_row:
                cell.border = Border(left=Side(border_style=cell.border.left.style), right=Side(border_style=cell.border.right.style), 
                                top=Side(border_style=cell.border.top.style), bottom=Side(border_style='thick'))


def add_it_ams_roles_sheet(wb):
    wb.create_sheet('IT_AMS_Roles')
    ws = wb['IT_AMS_Roles']
    alignment = Alignment(wrap_text=True, vertical='top', horizontal='left')
    center = Alignment(wrap_text=True, vertical='top', horizontal='center')
    border = Border(left=Side(border_style='thin'), right=Side(border_style='thin'), 
                                top=Side(border_style='thin'), bottom=Side(border_style='thin'))
    bold_font= Font(name='Arial', size=11, bold=True)
    italic_font= Font(name='Arial', size=10, italic=True)
    font= Font(name='Arial', size=10)
    fill = PatternFill('solid', fgColor='CCFFCC') #light green

    for row in ws['A1:D6']:
        for cell in row:
            cell.alignment = alignment
            cell.border = border

    ws['A1'] = 'IT_AMS_Roles'
    ws['A1'].font = bold_font
    ws['A1'].fill = fill
    ws['A2'] = 'PS'
    ws['A2'].font = bold_font
    ws['A2'].alignment = center
    ws['A3'] = 'GS'
    ws['A3'].font = bold_font
    ws['A3'].alignment = center
    ws['A4'] = 'PS and GS'
    ws['A4'].font = bold_font
    ws['A4'].alignment = center
    ws['A5'] = 'SPS'
    ws['A5'].font = bold_font
    ws['A5'].alignment = center
    ws['A6'] = 'RPM'
    ws['A6'].font = bold_font
    ws['A6'].alignment = center
    ws.column_dimensions['A'].width = 15

    ws['B1'] = 'Definitions'
    ws['B1'].font = bold_font
    ws['B1'].fill = fill
    ws['B2'] =  """When assigned on a RAN/Special/AIAN-Def/Follow-up review, the users with this role:
  1. Participates in the data collection, validation process when review is led by Regional office
  2. Supports Follow-up lead (FUL) in the data collection and validation process when review is led by DLH FUL
  3. Helps Regional Program Manager’s (RPM) or Follow-up Manager (FUM) to finalize the report"""
    ws['B2'].font = font
    ws['B3'] = """When assigned on a RAN/Special/AIAN-Def/Follow-Up review(s), the users with this role:
  1. Participates in the data collection process of the assigned review"""
    ws['B2'].font = font
    ws['B4'] = 'See above PS and GS access'
    ws['B2'].font = italic_font
    ws['B5'] = """When assigned on a RAN/Special/AIAN-Def/Follow-Up review(s), the users with this role:
  1. Participates in the report review process of the assigned review"""
    ws['B2'].font = font
    ws['B6'] = """When assigned on a RAN/Special/AIAN-Def/Follow-Up review(s), the users with this role:
  1. Participates in the report review process of the assigned review"""
    ws['B2'].font = font
    ws.column_dimensions['B'].width = 45

    ws['C1'] = 'IT-AMS Access'
    ws['C1'].font = bold_font
    ws['C1'].fill = fill
    ws['C2'] = """1. Home tab and its contents
2. Reviews tab and assigned reviews
3. Tasks tab and assigned tasks
4. Reports tab, assigned reports and Signed/shipped report of the assigned region
5. Grantees tab and assigned regions grantees monitoring history information
6. Dashboard tab and contents user is authorized to view"""
    ws['C2'].font = font
    ws['C3'] = """1. Home tab and its contents
2. Reviews tab and assigned reviews
3. Tasks tab and assigned tasks
4. Grantees tab and assigned regions grantees monitoring history information
5. Dashboard tab and contents user is authorized to view"""
    ws['C3'].font = font
    ws['C4'] = 'See above PS and GS access'
    ws['C4'].font = italic_font
    ws['C5'] = """1. Home tab and its contents
2. Reviews tab and assigned reviews
3. Tasks tab and assigned tasks
4. Reports tab, assigned reports and Signed/shipped report of the assigned region
5. Grantees tab and all regions grantees monitoring history information
6. My Regional reviews tab and reviews of the assigned region
7. Dashboard tab and contents user is authorized to view"""
    ws['C5'].font = font
    ws['C6'] = """1. Home tab and its contents
2. Reviews tab and assigned reviews
3. Tasks tab and assigned tasks
4. Reports tab, assigned reports and Signed/shipped report of the assigned region
5. Grantees tab and all regions grantees monitoring history information
6. My Regional reviews tab and reviews of the assigned region
7. Dashboard tab and contents user is authorized to view"""
    ws['C6'].font = font
    ws.column_dimensions['C'].width = 38

    ws['D1'] = 'Features Access'
    ws['D1'].font = bold_font
    ws['D1'].fill = fill
    ws['D2'] = """1. View and edit Review details of the assigned page
2. View and Edit data collection forms for assigned review
3. Read only access to Manifest, Eligibility files, Report preview, Pre-site, Evidence Binder, Findings page, grantee detail page, signed report search for all grantees across regions, all completed reviews across regions
4. View and share the Report with internal team/external users
5. View internal report shared log and reply to internal conversation thread
6. View the external Share report log
7. View and Edit the users self-profile"""
    ws['D2'].font = font
    ws['D3'] = """1. View Review details page of the assigned page
2. View and Edit data collection forms for assigned review
3. Read only access to Manifest, Eligibility files, Report preview, Pre-site, Evidence Binder, Findings page, grantee detail page, signed report search for all grantees across regions, all completed reviews across regions
4. View and Edit the user self-profile"""
    ws['D3'].font = font
    ws['D4'] = 'See above PS and GS access'
    ws['D4'].font = italic_font
    ws['D5'] = """1. View and edit Review details of the assigned page
2. View and Edit data collection forms for assigned review
3. Read only access to Manifest, Eligibility files, Report preview, Pre-site, Evidence Binder, Findings page, grantee detail page, signed report search for all grantees across regions, all completed reviews across regions
4. View and share the Report with internal team/external users
5. View internal report shared log and reply to internal conversation thread
6. View the external Share report log
7. View and Edit the user self-profile
8. View the list of reviews in their assigned region"""
    ws['D5'].font = font
    ws['D6'] = """1. View and edit Review details of the assigned page
2. View and Edit data collection forms for assigned review
3. Read only access to Manifest, Eligibility files, Report preview, Pre-site, Evidence Binder, Findings page, grantee detail page, signed report search for all grantees across regions, all completed reviews across regions
4. View and share the Report with internal team/external users
5. View internal report shared log and reply to internal conversation thread
6. View the external Share report log
7. View and Edit the user self-profile
8. View the list of reviews in their assigned region"""
    ws['D6'].font = font
    ws.column_dimensions['D'].width = 60


def highlight_reviewer_accounts_with_no_id_yellow(ws):
    col = ws.max_column
    for row in range(2, ws.max_row + 1):
        cell = ws.cell(row = row, column = col)
        if cell.value is None:
            cell.fill = PatternFill('solid', fgColor='FFFF00') #color yellow


def separate_title_groups_with_thick_borders(ws):
    col = 6 # the Title column
    for row in range(2, ws.max_row + 1):
        curr_title = ws.cell(row = row, column = col)
        next_title = ws.cell(row = row + 1, column = col)
        if curr_title.value != next_title.value:
            for column in range(1, ws.max_column + 1):
                cell = ws.cell(row = row, column = column)
                cell.border = Border(left=Side(border_style=cell.border.left.style), right=Side(border_style=cell.border.right.style), 
                                top=Side(border_style=cell.border.top.style), bottom=Side(border_style='thick'))


def separate_location_groups_with_thick_borders(ws):
    col = 2 # the Location column
    for row in range(2, ws.max_row + 1):
        curr_location = ws.cell(row = row, column = col)
        next_location = ws.cell(row = row + 1, column = col)
        if curr_location.value != next_location.value:
            for column in range(1, ws.max_column + 1):
                cell = ws.cell(row = row, column = column)
                cell.border = Border(left=Side(border_style=cell.border.left.style), right=Side(border_style=cell.border.right.style), 
                                top=Side(border_style=cell.border.top.style), bottom=Side(border_style='thick'))


def get_month_and_year():
    month = None
    year = None
    try:
        year = sys.argv[3]
        year = datetime.strptime(year, '%Y')
        year = year.strftime('%Y')
    except IndexError:
        year = None
    except ValueError:
        year = None
        print(f'INFO: Invalid year provided, will default accordingly')

    try:
        month = sys.argv[2]
        month = datetime.strptime(month, '%b')
        month = month.strftime('%b')
        year = date.today().strftime('%Y') if not year else year
    except IndexError:
        month = date.today() + relativedelta(months=1)
        print(f'INFO: Defaulting to the coming month ({month.strftime("%b")})')
        year = month.strftime('%Y') if not year else year
        month = month.strftime('%b')
    except ValueError:
        try:
            month = datetime.strptime(month, '%B')
            month = month.strftime('%b')
            year = date.today().strftime('%Y') if not year else year
        except ValueError:
            month = date.today() + relativedelta(months=1)
            print(f'INFO: Invalid month provided, will default to the coming month ({month.strftime("%b")})')
            year = month.strftime('%Y') if not year else year
            month = month.strftime('%b')
    
    return f'{month}-{year}'


if __name__ == '__main__':
    main()